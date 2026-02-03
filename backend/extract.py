from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import subprocess
import pathlib
import time

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"

def extract_doc(archivo):
    archivo = pathlib.Path(archivo)
    outdir = archivo.parent

    # 1️⃣ Convertir
    subprocess.run([
        SOFFICE,
        "--headless",
        "--convert-to", "txt:Text",
        str(archivo),
        "--outdir", str(outdir)
    ], check=True)

    # 2️⃣ Esperar un pelín (Windows + LO lo necesita)
    time.sleep(0.5)

    # 3️⃣ Buscar el txt generado
    txt_files = list(outdir.glob("*.txt"))

    if not txt_files:
        raise FileNotFoundError("LibreOffice no generó ningún .txt")

    # el más reciente (el recién convertido)
    txt_path = max(txt_files, key=lambda p: p.stat().st_mtime)

    # 4️⃣ Leer contenido
    return txt_path.read_text(encoding="utf-8", errors="ignore")


def extract_text_from_file(file_path, extension):
    try:
        if extension == "pdf":
            reader = PdfReader(file_path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        elif extension == "docx":
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)
        
        elif extension == "doc":
            doc = extract_doc(file_path)
            return doc

        elif extension in ("xls", "xlsx"):
            wb = load_workbook(file_path, data_only=True)
            text = []
            for sheet in wb:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text.append(str(cell.value))
            return "\n".join(text)

        elif extension in ("png", "jpg", "jpeg"):
            img = Image.open(file_path)
            return pytesseract.image_to_string(img)

        else:
            return ""

    except Exception as e:
        return ""
